"""Reading the files a construction site actually produces.

Site exports are Excel far more often than CSV, and they rarely put the header on the
first row: a punch list opens with a title like "RS01消项计划表2026/3/11" and the column
names sit underneath. A reader that assumes `csv.DictReader` semantics rejects the real
file and the operator is told to convert it by hand — which is how an ingestion pipeline
ends up unused.

This turns any of .csv/.xlsx/.xls into rows of {header: value}, finding the header row
rather than assuming it.
"""

from __future__ import annotations

import csv
import io
from typing import Any

#: How far into a sheet to look for the header. Site templates put a title, sometimes a
#: logo row, and occasionally a merged sub-title above it.
HEADER_SEARCH_ROWS = 12


class UnreadableFile(ValueError):
    """The file cannot be parsed, phrased so the uploader can act on it."""


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in ("none", "nan") else text


def _is_header_row(cells: list[str]) -> bool:
    """A header row has at least a couple of short, non-numeric labels.

    Two is the floor rather than three: a perfectly ordinary export can be two columns
    wide, and rejecting it would be a worse failure than occasionally reading a title
    row as a header. A one-cell row — which is what a sheet title looks like — never
    qualifies.
    """
    filled = [cell for cell in cells if cell]
    if len(filled) < 2:
        return False
    numeric = sum(1 for cell in filled if cell.replace(".", "", 1).replace("%", "").isdigit())
    if numeric > len(filled) / 2:
        return False
    return sum(len(cell) for cell in filled) / len(filled) <= 24


def _rows_to_records(rows: list[list[str]], source: str) -> tuple[list[dict[str, str]], list[str], int]:
    """Locate the header row, then read everything under it."""
    header_index = -1
    for index, cells in enumerate(rows[:HEADER_SEARCH_ROWS]):
        if _is_header_row(cells):
            header_index = index
            break
    if header_index < 0:
        raise UnreadableFile(
            f"No header row was found in the first {HEADER_SEARCH_ROWS} rows of {source}. "
            "The importer needs a row of column names above the data."
        )

    raw_header = rows[header_index]
    header: list[str] = []
    for position, cell in enumerate(raw_header):
        # Excel exports leave merged and blank header cells; they still need a stable key.
        header.append(cell.replace("\n", "").strip() or f"column_{position + 1}")

    records: list[dict[str, str]] = []
    for cells in rows[header_index + 1:]:
        if not any(cells):
            continue
        record = {header[i]: cells[i] for i in range(min(len(header), len(cells)))}
        if any(value for value in record.values()):
            records.append(record)
    return records, header, header_index + 1


def read_table(raw: bytes, filename: str) -> dict[str, Any]:
    """Return {"records", "header", "header_row", "format", "sheet"}."""
    name = (filename or "").lower()

    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        try:
            import openpyxl
        except ImportError:  # pragma: no cover - dependency is declared
            raise UnreadableFile("Excel support is not installed in this deployment (openpyxl).")
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        except Exception as exc:
            raise UnreadableFile(f"This file could not be opened as an Excel workbook: {exc}")
        sheet = workbook[workbook.sheetnames[0]]
        rows = [[_clean(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
        records, header, header_row = _rows_to_records(rows, f"sheet '{sheet.title}'")
        return {"records": records, "header": header, "header_row": header_row,
                "format": "xlsx", "sheet": sheet.title, "title_rows": [r for r in rows[:header_row - 1] if any(r)]}

    if name.endswith(".xls"):
        try:
            import xlrd
        except ImportError:  # pragma: no cover
            raise UnreadableFile("Legacy Excel support is not installed in this deployment (xlrd).")
        try:
            book = xlrd.open_workbook(file_contents=raw)
        except Exception as exc:
            raise UnreadableFile(f"This file could not be opened as a legacy Excel workbook: {exc}")
        sheet = book.sheet_by_index(0)
        rows = [[_clean(sheet.cell_value(r, c)) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
        records, header, header_row = _rows_to_records(rows, f"sheet '{sheet.name}'")
        return {"records": records, "header": header, "header_row": header_row,
                "format": "xls", "sheet": sheet.name, "title_rows": [r for r in rows[:header_row - 1] if any(r)]}

    text = raw.decode("utf-8-sig", errors="replace")
    rows = [[_clean(cell) for cell in row] for row in csv.reader(io.StringIO(text))]
    if not rows:
        raise UnreadableFile("The file is empty.")
    records, header, header_row = _rows_to_records(rows, "the CSV")
    return {"records": records, "header": header, "header_row": header_row,
            "format": "csv", "sheet": None, "title_rows": [r for r in rows[:header_row - 1] if any(r)]}
